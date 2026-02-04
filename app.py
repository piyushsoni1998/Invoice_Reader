"""
Invoice PDF OCR - HYBRID APPROACH
==================================
Simple Workflow:
1. Check each PDF page - is it IMAGE or TEXT?
2. IMAGE pages → Enhance resolution → Claude Vision (AI understanding)
3. TEXT pages → PaddleOCR (Local, FREE, Fast)

Requirements:
pip install fastapi uvicorn boto3 python-multipart pdf2image Pillow openpyxl numpy opencv-python-headless paddleocr paddlepaddle
"""

from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
import boto3
import json
import io
import base64
from typing import List, Dict, Any, Optional
import time
import re
from datetime import datetime

# ╔════════════════════════════════════════════════════════════════╗
# ║                    CONFIGURATION                                ║
# ╚════════════════════════════════════════════════════════════════╝

AWS_REGION = "us-east-1"
APP_PORT = 5000

# Claude Models
CLAUDE_HAIKU = "anthropic.claude-3-haiku-20240307-v1:0"
CLAUDE_SONNET = "anthropic.claude-3-sonnet-20240229-v1:0"

# Image Settings
DPI = 300
TARGET_RESOLUTION = 2000  # Target width/height for enhanced images
JPEG_QUALITY = 95
BATCH_SIZE = 3

# Enhancement Settings
CONTRAST_FACTOR = 1.3
SHARPNESS_FACTOR = 1.5
BRIGHTNESS_FACTOR = 1.1

# AWS Client
bedrock = boto3.client("bedrock-runtime", region_name=AWS_REGION)

# Initialize PaddleOCR (check availability at startup)
paddle_ocr = None
PADDLEOCR_AVAILABLE = False

# Check PaddleOCR availability once at startup
try:
    from paddleocr import PaddleOCR
    PADDLEOCR_AVAILABLE = True
    print("✅ PaddleOCR is available")
except ImportError:
    PADDLEOCR_AVAILABLE = False
    print("⚠️ PaddleOCR not installed - will use Claude for all pages")
except Exception as e:
    PADDLEOCR_AVAILABLE = False
    print(f"⚠️ PaddleOCR error: {e}")

def get_paddle_ocr():
    """Get or initialize PaddleOCR instance"""
    global paddle_ocr
    if not PADDLEOCR_AVAILABLE:
        return None
    if paddle_ocr is None:
        try:
            paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)
            print("   ✅ PaddleOCR initialized")
        except Exception as e:
            print(f"   ⚠️ PaddleOCR init error: {e}")
            return None
    return paddle_ocr

app = FastAPI(title="Invoice OCR - Hybrid Approach")

# Store results for download
results_store: Dict[str, Any] = {}


# ═══════════════════════════════════════════════════════════════════
# PAGE TYPE DETECTION - Simple: IMAGE or TEXT?
# ═══════════════════════════════════════════════════════════════════

def is_image_page(pdf_bytes: bytes, page_num: int) -> Dict[str, Any]:
    """
    Check if a PDF page is an IMAGE (scanned) or TEXT (digital).

    Returns:
        {
            'is_image': True/False,
            'has_text': True/False,
            'text_length': int,
            'image_count': int,
            'reason': str
        }
    """
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        page = doc[page_num]

        # Get extractable text
        text = page.get_text("text").strip()
        text_length = len(text)

        # Get images in the page
        images = page.get_images()
        image_count = len(images)

        # Get page dimensions
        page_rect = page.rect
        page_area = page_rect.width * page_rect.height

        # Check if images cover most of the page
        total_image_area = 0
        for img in images:
            xref = img[0]
            try:
                img_rect = page.get_image_bbox(img)
                if img_rect:
                    img_area = img_rect.width * img_rect.height
                    total_image_area += img_area
            except:
                pass

        image_coverage = total_image_area / page_area if page_area > 0 else 0

        doc.close()

        # Decision Logic:
        # - If very little text AND has images covering >50% → IMAGE page
        # - If has substantial text → TEXT page

        if text_length < 50 and image_count > 0:
            return {
                'is_image': True,
                'has_text': False,
                'text_length': text_length,
                'image_count': image_count,
                'image_coverage': image_coverage,
                'reason': 'No extractable text, has images'
            }
        elif text_length < 100 and image_coverage > 0.5:
            return {
                'is_image': True,
                'has_text': False,
                'text_length': text_length,
                'image_count': image_count,
                'image_coverage': image_coverage,
                'reason': 'Images cover most of page'
            }
        else:
            return {
                'is_image': False,
                'has_text': True,
                'text_length': text_length,
                'image_count': image_count,
                'image_coverage': image_coverage,
                'reason': 'Has extractable text'
            }

    except ImportError:
        # PyMuPDF not available, assume all pages are images
        return {
            'is_image': True,
            'has_text': False,
            'text_length': 0,
            'image_count': 0,
            'image_coverage': 0,
            'reason': 'PyMuPDF not available'
        }
    except Exception as e:
        return {
            'is_image': True,
            'has_text': False,
            'text_length': 0,
            'image_count': 0,
            'image_coverage': 0,
            'reason': f'Error: {str(e)}'
        }


# ═══════════════════════════════════════════════════════════════════
# IMAGE ENHANCEMENT - Increase Resolution
# ═══════════════════════════════════════════════════════════════════

def enhance_image_resolution(img, target_size: int = TARGET_RESOLUTION):
    """
    Enhance image resolution for better OCR.
    - Upscale if too small
    - Enhance contrast, sharpness
    - Denoise
    """
    from PIL import Image, ImageEnhance, ImageFilter, ImageOps

    try:
        # Convert to RGB
        if img.mode != 'RGB':
            img = img.convert('RGB')

        original_size = max(img.width, img.height)

        # 1. Upscale if image is too small
        if original_size < target_size:
            scale_factor = target_size / original_size
            new_width = int(img.width * scale_factor)
            new_height = int(img.height * scale_factor)
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            print(f"(upscaled {scale_factor:.1f}x)", end=" ")

        # 2. Auto-contrast
        img = ImageOps.autocontrast(img, cutoff=1)

        # 3. Enhance contrast
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(CONTRAST_FACTOR)

        # 4. Enhance sharpness
        enhancer = ImageEnhance.Sharpness(img)
        img = enhancer.enhance(SHARPNESS_FACTOR)

        # 5. Slight brightness boost
        enhancer = ImageEnhance.Brightness(img)
        img = enhancer.enhance(BRIGHTNESS_FACTOR)

        # 6. Unsharp mask for edge enhancement
        img = img.filter(ImageFilter.UnsharpMask(radius=2, percent=150, threshold=3))

        return img

    except Exception as e:
        print(f"⚠️ Enhancement failed: {e}")
        return img


def check_and_enhance_resolution(img) -> tuple:
    """
    Check image resolution and enhance if needed.
    Returns (enhanced_img, resolution_info)
    """
    original_width = img.width
    original_height = img.height
    original_size = max(original_width, original_height)

    # Determine if enhancement needed
    needs_enhancement = original_size < TARGET_RESOLUTION

    if needs_enhancement:
        enhanced_img = enhance_image_resolution(img, TARGET_RESOLUTION)
        resolution_info = {
            'original': f"{original_width}x{original_height}",
            'enhanced': f"{enhanced_img.width}x{enhanced_img.height}",
            'was_enhanced': True
        }
    else:
        # Just apply basic enhancement without upscaling
        enhanced_img = enhance_image_resolution(img, original_size)
        resolution_info = {
            'original': f"{original_width}x{original_height}",
            'enhanced': f"{enhanced_img.width}x{enhanced_img.height}",
            'was_enhanced': False
        }

    return enhanced_img, resolution_info


# ═══════════════════════════════════════════════════════════════════
# PADDLEOCR - For TEXT Pages (Local, FREE)
# ═══════════════════════════════════════════════════════════════════

def extract_text_with_paddle(img) -> str:
    """
    Extract text from image using PaddleOCR.
    This is FREE and runs locally!
    """
    ocr = get_paddle_ocr()
    if ocr is None:
        return "[PaddleOCR not available]"

    try:
        import numpy as np

        # Convert PIL to numpy array
        img_array = np.array(img)

        # Run OCR
        result = ocr.ocr(img_array, cls=True)

        # Extract text from results
        text_lines = []
        if result and result[0]:
            for line in result[0]:
                if line and len(line) >= 2:
                    text = line[1][0]  # Get text content
                    confidence = line[1][1]  # Get confidence
                    text_lines.append(text)

        return "\n".join(text_lines)

    except Exception as e:
        return f"[PaddleOCR Error: {str(e)}]"


# ═══════════════════════════════════════════════════════════════════
# CLAUDE VISION - For IMAGE Pages
# ═══════════════════════════════════════════════════════════════════

CLAUDE_SYSTEM = """You are an OCR tool that reads and transcribes ALL text from images.

Your job:
- Read ALL visible text in the image
- Output the text exactly as written
- Include numbers, dates, names, addresses, amounts
- Include handwritten text
- Include table contents
- Include fine print

Output format: Just transcribe the text, organized by section."""

CLAUDE_PROMPT = """Please transcribe ALL text visible in this image.

Include:
- All printed text
- All handwritten text
- Numbers, dates, amounts
- Names, addresses
- Table contents
- Any other visible text

BEGIN TRANSCRIPTION:"""


def extract_text_with_claude(img, model_id: str = CLAUDE_HAIKU) -> tuple:
    """
    Extract text from image using Claude Vision.
    Used for IMAGE pages (scanned documents, photos, handwriting).
    """
    try:
        # Convert to base64
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=JPEG_QUALITY, optimize=True)
        img_bytes = buffer.getvalue()

        # Check size and reduce if needed
        if len(img_bytes) > 4.5 * 1024 * 1024:
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=75, optimize=True)
            img_bytes = buffer.getvalue()

        img_base64 = base64.b64encode(img_bytes).decode('utf-8')

        # Build request
        content = [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": img_base64
                }
            },
            {"type": "text", "text": CLAUDE_PROMPT}
        ]

        body = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 4096,
            "temperature": 0.1,
            "system": CLAUDE_SYSTEM,
            "messages": [{"role": "user", "content": content}]
        }

        response = bedrock.invoke_model(
            modelId=model_id,
            body=json.dumps(body),
            contentType="application/json",
            accept="application/json"
        )

        result = json.loads(response["body"].read())
        text = result["content"][0]["text"]

        model_name = "Haiku" if "haiku" in model_id else "Sonnet"
        return text, model_name, True

    except Exception as e:
        return f"[Claude Error: {str(e)}]", "Error", False


# ═══════════════════════════════════════════════════════════════════
# MAIN PROCESSING FUNCTION
# ═══════════════════════════════════════════════════════════════════

def process_pdf_hybrid(pdf_bytes: bytes, enhance: bool = True) -> Dict[str, Any]:
    """
    Process PDF with hybrid approach:
    - IMAGE pages → Claude Vision
    - TEXT pages → PaddleOCR
    """
    from pdf2image import convert_from_bytes

    result = {
        'pages': [],
        'stats': {
            'total_pages': 0,
            'image_pages': 0,
            'text_pages': 0,
            'claude_calls': 0,
            'paddle_calls': 0
        },
        'combined_text': ''
    }

    # Convert PDF to images
    print(f"   📄 Converting PDF to images at {DPI} DPI...")
    images = convert_from_bytes(pdf_bytes, dpi=DPI, fmt='PNG')
    result['stats']['total_pages'] = len(images)

    all_text = []

    for idx, img in enumerate(images):
        page_num = idx + 1
        print(f"\n   📄 Page {page_num}/{len(images)}: ", end="")

        # Step 1: Check if page is IMAGE or TEXT
        page_info = is_image_page(pdf_bytes, idx)

        page_result = {
            'page_num': page_num,
            'is_image': page_info['is_image'],
            'reason': page_info['reason'],
            'method': '',
            'text': '',
            'resolution': {}
        }

        if page_info['is_image']:
            # IMAGE PAGE → Claude Vision
            print(f"🖼️ IMAGE ({page_info['reason']}) → ", end="")

            # Step 2: Check and enhance resolution
            if enhance:
                enhanced_img, res_info = check_and_enhance_resolution(img)
                page_result['resolution'] = res_info
            else:
                enhanced_img = img
                page_result['resolution'] = {
                    'original': f"{img.width}x{img.height}",
                    'was_enhanced': False
                }

            # Step 3: Send to Claude
            print("Claude... ", end="")
            text, model_used, success = extract_text_with_claude(enhanced_img, CLAUDE_HAIKU)

            if not success or "error" in text.lower():
                # Try Sonnet as fallback
                print("(trying Sonnet)... ", end="")
                text, model_used, success = extract_text_with_claude(enhanced_img, CLAUDE_SONNET)

            page_result['method'] = f"Claude {model_used}"
            page_result['text'] = text
            result['stats']['image_pages'] += 1
            result['stats']['claude_calls'] += 1
            print(f"✅ {model_used}")

        else:
            # TEXT PAGE → PaddleOCR
            print(f"📝 TEXT ({page_info['text_length']} chars) → PaddleOCR... ", end="")

            # Basic enhancement for OCR
            if enhance:
                enhanced_img, res_info = check_and_enhance_resolution(img)
                page_result['resolution'] = res_info
            else:
                enhanced_img = img

            text = extract_text_with_paddle(enhanced_img)

            page_result['method'] = "PaddleOCR"
            page_result['text'] = text
            result['stats']['text_pages'] += 1
            result['stats']['paddle_calls'] += 1
            print("✅")

        result['pages'].append(page_result)

        # Add to combined text
        header = f"\n{'='*60}\n📄 PAGE {page_num} [{page_result['method']}]\n{'='*60}\n"
        all_text.append(header + page_result['text'])

    result['combined_text'] = "\n\n".join(all_text)

    # Print summary
    print(f"\n   📊 Summary:")
    print(f"      Total pages: {result['stats']['total_pages']}")
    print(f"      Image pages (Claude): {result['stats']['image_pages']}")
    print(f"      Text pages (PaddleOCR): {result['stats']['text_pages']}")

    return result


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════

def create_excel_from_data(extracted_data: Dict[str, Any], filename: str) -> bytes:
    """Create Excel file from extracted data"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Extraction Report"

    ws['A1'] = "Invoice OCR Extraction Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Source: {filename}"

    # Page summary table
    row = 5
    headers = ['Page', 'Type', 'Method', 'Resolution', 'Text Preview']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for page in extracted_data.get('pages', []):
        row += 1
        res = page.get('resolution', {})
        res_str = res.get('original', 'N/A')
        if res.get('was_enhanced'):
            res_str += f" → {res.get('enhanced', '')}"

        values = [
            page.get('page_num', ''),
            'Image' if page.get('is_image') else 'Text',
            page.get('method', ''),
            res_str,
            page.get('text', '')[:100] + '...' if len(page.get('text', '')) > 100 else page.get('text', '')
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(value))
            cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 60

    # Sheet 2: Full Text
    ws2 = wb.create_sheet("Full Text")
    ws2['A1'] = "Complete Extracted Text"
    ws2['A1'].font = Font(bold=True, size=14)

    full_text = extracted_data.get('combined_text', '')
    lines = full_text.split('\n')
    for idx, line in enumerate(lines, start=3):
        ws2.cell(row=idx, column=1, value=line)

    ws2.column_dimensions['A'].width = 120

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
def home():
    # Use startup check result - do NOT import PaddleOCR here (causes reinitialization error)
    if PADDLEOCR_AVAILABLE:
        paddle_status = '<span class="badge badge-green">✓ PaddleOCR</span>'
    else:
        paddle_status = '<span class="badge badge-red">✗ PaddleOCR</span>'

    return f"""
<!DOCTYPE html>
<html>
<head>
    <title>Invoice OCR - Hybrid</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            background: linear-gradient(135deg, #1565c0 0%, #0d47a1 100%);
            min-height: 100vh;
            padding: 40px 20px;
        }}
        .container {{ max-width: 800px; margin: 0 auto; }}
        .card {{
            background: white;
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }}
        h1 {{ color: #1565c0; margin-bottom: 8px; }}
        .badges {{ margin-bottom: 20px; }}
        .badge {{
            display: inline-block;
            padding: 5px 12px;
            border-radius: 20px;
            font-size: 12px;
            margin-right: 8px;
            margin-bottom: 5px;
            color: white;
        }}
        .badge-blue {{ background: #1565c0; }}
        .badge-green {{ background: #2e7d32; }}
        .badge-orange {{ background: #e65100; }}
        .badge-red {{ background: #c62828; }}

        .workflow {{
            background: #f5f5f5;
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
        }}
        .workflow h3 {{ color: #1565c0; margin-bottom: 15px; }}
        .workflow-step {{
            display: flex;
            align-items: center;
            padding: 10px;
            margin: 8px 0;
            background: white;
            border-radius: 8px;
            border-left: 4px solid #1565c0;
        }}
        .step-icon {{ font-size: 24px; margin-right: 15px; }}
        .step-text strong {{ color: #1565c0; }}
        .step-arrow {{ color: #1565c0; font-size: 20px; margin: 0 10px; }}

        .upload-area {{
            border: 3px dashed #1565c0;
            border-radius: 12px;
            padding: 40px 30px;
            text-align: center;
            background: #e3f2fd;
            transition: all 0.3s;
            cursor: pointer;
        }}
        .upload-area:hover {{ border-color: #0d47a1; background: #bbdefb; }}
        input[type="file"] {{ margin: 15px 0; font-size: 14px; }}

        .options {{ margin: 20px 0; padding: 15px; background: #f5f5f5; border-radius: 8px; }}
        .options label {{ display: flex; align-items: center; margin: 8px 0; cursor: pointer; }}
        .options input[type="checkbox"] {{ margin-right: 10px; width: 18px; height: 18px; }}

        button {{
            background: linear-gradient(135deg, #1565c0 0%, #1976d2 100%);
            color: white;
            border: none;
            padding: 15px 40px;
            font-size: 16px;
            border-radius: 8px;
            cursor: pointer;
        }}
        button:hover {{ transform: scale(1.02); }}

        .loading {{ display: none; text-align: center; padding: 40px; }}
        .spinner {{
            width: 50px; height: 50px;
            border: 4px solid #f3f3f3;
            border-top: 4px solid #1565c0;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }}
        @keyframes spin {{ 0% {{ transform: rotate(0deg); }} 100% {{ transform: rotate(360deg); }} }}
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <h1>📄 Invoice OCR - Hybrid</h1>
            <div class="badges">
                <span class="badge badge-blue">🤖 Claude Vision</span>
                {paddle_status}
                <span class="badge badge-orange">🔧 Auto-Enhance</span>
            </div>
            <p style="color: #666; margin-bottom: 20px;">Simple & Smart: Right tool for each page type</p>

            <div class="workflow">
                <h3>📋 How It Works</h3>

                <div class="workflow-step">
                    <span class="step-icon">📄</span>
                    <div class="step-text">
                        <strong>Step 1:</strong> Check each PDF page - is it IMAGE or TEXT?
                    </div>
                </div>

                <div class="workflow-step">
                    <span class="step-icon">🖼️</span>
                    <div class="step-text">
                        <strong>IMAGE pages</strong> (scanned, photos, handwriting)
                        <span class="step-arrow">→</span>
                        Enhance Resolution
                        <span class="step-arrow">→</span>
                        <strong>Claude Vision</strong>
                    </div>
                </div>

                <div class="workflow-step">
                    <span class="step-icon">📝</span>
                    <div class="step-text">
                        <strong>TEXT pages</strong> (digital PDFs)
                        <span class="step-arrow">→</span>
                        <strong>PaddleOCR</strong> (FREE, Local)
                    </div>
                </div>
            </div>

            <form id="uploadForm" action="/upload" method="post" enctype="multipart/form-data">
                <div class="upload-area" id="uploadArea">
                    <div style="font-size: 40px; margin-bottom: 10px;">📄</div>
                    <p style="font-size: 16px; margin-bottom: 10px;">Upload PDF, image, or document</p>
                    <input type="file" name="file" accept=".pdf,.png,.jpg,.jpeg" required>

                    <div class="options">
                        <label><input type="checkbox" name="enhance" checked> 🔧 Enhance image resolution (recommended)</label>
                    </div>

                    <button type="submit">🚀 Process Document</button>
                </div>
            </form>

            <div class="loading" id="loadingDiv">
                <div class="spinner"></div>
                <p style="font-size: 18px; color: #1565c0;">Processing document...</p>
                <p style="color: #666; font-size: 14px;">Analyzing pages and extracting text...</p>
            </div>
        </div>
    </div>

    <script>
        document.getElementById('uploadForm').addEventListener('submit', function() {{
            document.getElementById('uploadArea').style.display = 'none';
            document.getElementById('loadingDiv').style.display = 'block';
        }});
    </script>
</body>
</html>
"""


@app.post("/upload", response_class=HTMLResponse)
async def upload_file(
    file: UploadFile = File(...),
    enhance: bool = True
):
    """Process document with hybrid approach"""

    filename = file.filename
    file_bytes = await file.read()
    start_time = time.time()
    result_id = f"{int(time.time())}_{filename}"

    print(f"\n{'='*70}")
    print(f"📄 HYBRID OCR PROCESSING")
    print(f"📄 File: {filename}")
    print(f"🔧 Enhancement: {'ON' if enhance else 'OFF'}")
    print(f"{'='*70}")

    try:
        if filename.lower().endswith('.pdf'):
            # Process PDF with hybrid approach
            result = process_pdf_hybrid(file_bytes, enhance=enhance)

        else:
            # Single image - use Claude
            from PIL import Image
            img = Image.open(io.BytesIO(file_bytes))

            print(f"\n   🖼️ Single image → Claude Vision")

            if enhance:
                enhanced_img, res_info = check_and_enhance_resolution(img)
            else:
                enhanced_img = img
                res_info = {'original': f"{img.width}x{img.height}", 'was_enhanced': False}

            text, model_used, success = extract_text_with_claude(enhanced_img, CLAUDE_HAIKU)

            result = {
                'pages': [{
                    'page_num': 1,
                    'is_image': True,
                    'method': f"Claude {model_used}",
                    'text': text,
                    'resolution': res_info
                }],
                'stats': {
                    'total_pages': 1,
                    'image_pages': 1,
                    'text_pages': 0,
                    'claude_calls': 1,
                    'paddle_calls': 0
                },
                'combined_text': text
            }

        # Store results for download
        results_store[result_id] = {
            "filename": filename,
            "result": result,
            "timestamp": datetime.now().isoformat()
        }

        elapsed = time.time() - start_time
        print(f"\n✅ Complete in {elapsed:.1f}s")

        return render_result(filename, result, elapsed, result_id, enhance)

    except Exception as e:
        import traceback
        print(f"❌ Error: {e}")
        traceback.print_exc()
        return render_error("Processing Error", str(e))


@app.get("/download/excel/{result_id}")
async def download_excel(result_id: str):
    """Download results as Excel file"""
    if result_id not in results_store:
        return HTMLResponse("<h1>Result not found</h1>", status_code=404)

    data = results_store[result_id]
    excel_bytes = create_excel_from_data(data["result"], data["filename"])
    filename = data["filename"].rsplit('.', 1)[0]

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}_extracted.xlsx"}
    )


@app.get("/download/text/{result_id}")
async def download_text(result_id: str):
    """Download results as text file"""
    if result_id not in results_store:
        return HTMLResponse("<h1>Result not found</h1>", status_code=404)

    data = results_store[result_id]
    text_bytes = data["result"]["combined_text"].encode('utf-8')
    filename = data["filename"].rsplit('.', 1)[0]

    return StreamingResponse(
        io.BytesIO(text_bytes),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}_extracted.txt"}
    )


def render_result(filename: str, result: Dict, elapsed: float, result_id: str, enhanced: bool) -> str:
    def esc(s): return s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    stats = result['stats']

    # Build page breakdown
    page_rows = ""
    for page in result['pages']:
        res = page.get('resolution', {})
        res_str = res.get('original', 'N/A')
        if res.get('was_enhanced'):
            res_str += f" → {res.get('enhanced', '')}"

        type_badge = '🖼️ Image' if page.get('is_image') else '📝 Text'
        method_badge = page.get('method', 'Unknown')

        page_rows += f"""
        <tr>
            <td>{page.get('page_num', '')}</td>
            <td>{type_badge}</td>
            <td><strong>{method_badge}</strong></td>
            <td style="font-size: 11px;">{res_str}</td>
        </tr>
        """

    return f"""
<!DOCTYPE html>
<html>
<head>
    <title>Results - {filename}</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Arial; background: #f5f5f5; padding: 20px; margin: 0; }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        .header {{ background: linear-gradient(135deg, #1565c0, #1976d2); color: white; padding: 20px 25px; border-radius: 12px; margin-bottom: 20px; }}
        .stats {{ display: flex; gap: 15px; margin-top: 15px; flex-wrap: wrap; }}
        .stat {{ background: rgba(255,255,255,0.2); padding: 10px 20px; border-radius: 8px; text-align: center; }}
        .stat-val {{ font-size: 24px; font-weight: bold; }}
        .stat-lbl {{ font-size: 11px; opacity: 0.9; }}

        .content {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        @media (max-width: 1000px) {{ .content {{ grid-template-columns: 1fr; }} }}

        .panel {{ background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .panel h3 {{ color: #1565c0; margin-bottom: 15px; }}

        textarea {{ width: 100%; height: 450px; font-family: Consolas, monospace; font-size: 12px; padding: 15px; border: 1px solid #ddd; border-radius: 6px; background: #fafafa; resize: vertical; }}

        table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
        th {{ background: #1565c0; color: white; padding: 10px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f5f5f5; }}

        .actions {{ margin-top: 20px; display: flex; gap: 10px; flex-wrap: wrap; }}
        .btn {{ padding: 12px 20px; border-radius: 6px; text-decoration: none; cursor: pointer; border: none; font-size: 14px; display: inline-flex; align-items: center; gap: 8px; }}
        .btn-primary {{ background: #1565c0; color: white; }}
        .btn-green {{ background: #2e7d32; color: white; }}
        .btn-outline {{ background: white; color: #1565c0; border: 2px solid #1565c0; }}
        .btn:hover {{ opacity: 0.9; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>📄 Extraction Complete</h2>
            <p>{filename}</p>
            <div class="stats">
                <div class="stat"><div class="stat-val">{stats['total_pages']}</div><div class="stat-lbl">Total Pages</div></div>
                <div class="stat"><div class="stat-val">{elapsed:.1f}s</div><div class="stat-lbl">Time</div></div>
                <div class="stat"><div class="stat-val">{stats['image_pages']}</div><div class="stat-lbl">🖼️ Claude</div></div>
                <div class="stat"><div class="stat-val">{stats['text_pages']}</div><div class="stat-lbl">📝 PaddleOCR</div></div>
            </div>
        </div>

        <div class="content">
            <div class="panel">
                <h3>📋 Extracted Text</h3>
                <textarea readonly>{esc(result['combined_text'])}</textarea>
            </div>

            <div class="panel">
                <h3>📊 Page Breakdown</h3>
                <table>
                    <tr>
                        <th>Page</th>
                        <th>Type</th>
                        <th>Method</th>
                        <th>Resolution</th>
                    </tr>
                    {page_rows}
                </table>
            </div>
        </div>

        <div class="actions">
            <a href="/" class="btn btn-primary">⬅ Process Another</a>
            <a href="/download/excel/{result_id}" class="btn btn-green">📊 Download Excel</a>
            <a href="/download/text/{result_id}" class="btn btn-outline">📄 Download Text</a>
            <button class="btn btn-outline" onclick="navigator.clipboard.writeText(document.querySelector('textarea').value);alert('Copied!')">📋 Copy Text</button>
        </div>
    </div>
</body>
</html>
"""


def render_error(title: str, msg: str) -> str:
    return f"""
<!DOCTYPE html>
<html>
<head><title>Error</title>
<style>body {{ font-family: Arial; max-width: 600px; margin: 50px auto; padding: 20px; }}
.error {{ background: #ffebee; border-left: 4px solid #f44336; padding: 20px; border-radius: 0 8px 8px 0; }}
.btn {{ display: inline-block; background: #1565c0; color: white; padding: 12px 25px; text-decoration: none; border-radius: 6px; margin-top: 20px; }}</style>
</head>
<body>
<h2>❌ {title}</h2>
<div class="error"><pre>{msg}</pre></div>
<a href="/" class="btn">⬅ Try Again</a>
</body>
</html>
"""


if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*70)
    print("📄 INVOICE OCR - HYBRID APPROACH")
    print("="*70)
    print("Workflow:")
    print("  1. Check each page: IMAGE or TEXT?")
    print("  2. IMAGE pages → Enhance → Claude Vision")
    print("  3. TEXT pages → PaddleOCR (FREE, Local)")
    print("="*70)
    print(f"🚀 http://127.0.0.1:{APP_PORT}")
    print("="*70 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=APP_PORT)
